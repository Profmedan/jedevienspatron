#!/usr/bin/env python3
"""
Rebuild jeu_interactif_v2.xlsx with BLOCK_SIZE=67 and CR section.
Hardcoded version: Use simple hardcoded BEFORE values for turn 1.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_ROWS = 11
BLOCK_SIZE = 67
MAX_TURNS = 12

COLORS = {
    'dark_blue': '1F3864',
    'light_blue': '2E75B6',
    'orange': 'ED7D31',
    'light_gray': 'D9D9D9',
    'bilan_actif': 'DEEAF1',
    'bilan_passif': 'FCE4D6',
    'bilan_total': 'BDD7EE',
    'result_yellow': 'FFF2CC',
    'cr_produit': 'E2EFDA',
    'input_yellow': 'FFFF00',
}

VALIDATION_LISTS = {
    'enterprise': ['Manufacture Belvaux', 'Véloce Transports', 'Azura Commerce', 'Synergia Lab'],
    'trimestres': ['6', '8', '10', '12'],
    'emprunt': ['0', '5000', '8000', '12000', '16000', '20000'],
    'mode': ['Tréso', 'Dettes D+1'],
    'recr': ['Aucun', 'Commercial Junior', 'Commercial Senior', 'Directrice Commerciale'],
    'carte': ['Aucune', 'Camionnette', 'Berline', 'Site Internet', 'RSE', 'R&D', 'Expansion',
              'Certification ISO', 'Application Mobile', 'Prêt Bancaire', 'Levée de Fonds',
              'Crédit-Bail', 'Crowdfunding', 'Publicité', 'Relance Clients', 'Formation',
              'Programme Fidélité', 'Export International', 'Partenariat Commercial', 'Affacturage',
              'Contrat Maintenance', 'Assurance Prévoyance', 'Mutuelle Collective', 'Cybersécurité',
              'Fourgon Réfrigéré', 'Vélo Cargo', 'ERP', 'Marketplace', 'Entrepôt Automatisé',
              'Label Qualité', 'Achat d\'Urgence', 'Maintenance Préventive', 'Révision Générale',
              'Optimisation Lean', 'Sous-traitance'],
    'event': ['Aucun', 'Client VIP', 'Contrôle Fiscal', 'Subvention Innovation', 'Placement Financier',
              'Crise Sanitaire', 'Incendie', 'Grève', 'Bouche à Oreille', 'Perte de Données',
              'Développement Durable', 'Prix PME', 'Rupture Fournisseur', 'Litige Commercial'],
}

# BEFORE values for each enterprise (actif: immo, stocks, creances, treso; passif: capitaux, emprunts, dettes, decouvert, resultat)
ENTERPRISE_BEFORE = {
    'Manufacture Belvaux': {'immo': 8000, 'stocks': 4000, 'creances': 2000, 'treso': 2000,
                            'capitaux': 8000, 'emprunts': 8000, 'dettes': 0, 'decouvert': 0, 'resultat': 0},
    'Véloce Transports': {'immo': 10000, 'stocks': 2000, 'creances': 1000, 'treso': 3000,
                          'capitaux': 8000, 'emprunts': 8000, 'dettes': 0, 'decouvert': 0, 'resultat': 0},
    'Azura Commerce': {'immo': 4000, 'stocks': 8000, 'creances': 2000, 'treso': 2000,
                       'capitaux': 8000, 'emprunts': 8000, 'dettes': 0, 'decouvert': 0, 'resultat': 0},
    'Synergia Lab': {'immo': 6000, 'stocks': 2000, 'creances': 3000, 'treso': 2000,
                     'capitaux': 6500, 'emprunts': 6500, 'dettes': 0, 'decouvert': 0, 'resultat': 0},
}

def read_donnees_sheet(xlsx_path):
    try:
        wb = load_workbook(xlsx_path)
        sheet = wb['Données']
        data = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    data[(cell.row, cell.column)] = cell.value
        return data
    except Exception as e:
        print(f"ERROR reading Données: {e}")
        return None

def copy_donnees_to_new(new_wb, donnees_data):
    sheet = new_wb['Données']
    for (row, col), value in donnees_data.items():
        sheet.cell(row, col).value = value

def add_dv(sheet, cell_ref, formula):
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.add(cell_ref)
    sheet.add_data_validation(dv)

def format_cell(cell, bg_color=None, font_color=None, bold=False):
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    if font_color or bold:
        cell.font = Font(color=font_color or '000000', bold=bold, size=10, name='Arial')
    cell.alignment = Alignment(horizontal='center', vertical='center')

def set_number_format(cell):
    cell.number_format = '#,##0;-#,##0;"-"'

def build_simulation_sheet(ws):
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 7

    ws['A1'] = 'SIMULATION JEU INTERACTIF'
    ws['A1'].font = Font(bold=True, size=14, name='Arial')

    ws['A3'] = 'Entreprise :'
    ws['D3'] = 'Manufacture Belvaux'
    format_cell(ws['D3'], bg_color=COLORS['input_yellow'], font_color='0000FF')
    add_dv(ws, 'D3', '"' + ','.join(VALIDATION_LISTS['enterprise']) + '"')

    ws['A4'] = 'Nombre de trimestres :'
    ws['D4'] = 6
    format_cell(ws['D4'], bg_color=COLORS['input_yellow'], font_color='0000FF')
    add_dv(ws, 'D4', '"' + ','.join(VALIDATION_LISTS['trimestres']) + '"')

    ws['A5'] = 'Trimestre initial :'
    ws['D5'] = 1

    ws['A11'] = 'TRIMESTRE'
    ws['B11'] = 'AVANT'
    ws['C11'] = 'DÉCISIONS'
    ws['D11'] = 'MONTANT'
    ws['E11'] = 'APRÈS'
    ws['F11'] = 'NOTE'

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = ws[f'{col}11']
        format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    for turn_num in range(1, MAX_TURNS + 1):
        build_turn_block(ws, turn_num)

def build_turn_block(ws, turn_num):
    bs = HEADER_ROWS + 1 + (turn_num - 1) * BLOCK_SIZE

    r = bs
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = f'TRIMESTRE {turn_num}'
    format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    r = bs + 1
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = '📋 DÉCISIONS'
    format_cell(cell, bg_color=COLORS['light_blue'], font_color='FFFFFF', bold=True)

    decisions = [
        (3, 'Emprunt demandé :', 0, VALIDATION_LISTS['emprunt']),
        (4, 'Achat de stock :', 0, None),
        (5, 'Mode de trésorerie :', 'Tréso', VALIDATION_LISTS['mode']),
        (6, 'Recrutement :', 'Aucun', VALIDATION_LISTS['recr']),
        (7, 'Carte action :', 'Aucune', VALIDATION_LISTS['carte']),
        (8, 'Événement :', 'Aucun', VALIDATION_LISTS['event']),
    ]

    for offset, label, default_val, dv_list in decisions:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'A{r}'].alignment = Alignment(horizontal='left')
        ws[f'D{r}'] = default_val
        format_cell(ws[f'D{r}'], bg_color=COLORS['input_yellow'], font_color='0000FF')
        if isinstance(default_val, (int, float)):
            set_number_format(ws[f'D{r}'])
        if dv_list:
            add_dv(ws, f'D{r}', '"' + ','.join(dv_list) + '"')

    r = bs + 10
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = '⚙️ EFFETS AUTOMATIQUES'
    format_cell(cell, bg_color=COLORS['orange'], font_color='FFFFFF', bold=True)

    effets = [
        (11, 'Charges courantes'), (12, 'Amortissements'), (13, 'Remboursements'),
        (14, 'Agios'), (15, 'Règlement dettes'), (18, 'Stock acheté'), (19, 'Stock à payer'),
        (22, 'Capacité prod.'), (23, 'Stocks dispo'), (24, 'Unités vendues'),
        (25, 'CA (Chiffre affaires)'), (26, 'CMV (Coût matière)'),
        (29, 'Immobilisations'), (30, 'Trésorerie'), (31, 'Stocks'),
        (34, 'Trésorerie'), (35, 'Stocks'),
    ]

    for offset, label in effets:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'A{r}'].alignment = Alignment(horizontal='left')
        ws[f'D{r}'] = 0
        set_number_format(ws[f'D{r}'])

    r = bs + 39
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = '📊 BILAN'
    format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    r = bs + 40
    ws[f'A{r}'] = 'ACTIF'
    ws[f'C{r}'] = 'AVANT'
    ws[f'E{r}'] = 'APRÈS'
    for col in ['A', 'C', 'E']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['light_gray'], bold=True)

    actifs = [
        (40, 'Immobilisations', 'immo'),
        (41, 'Stocks', 'stocks'),
        (42, 'Créances clients', 'creances'),
        (43, 'Trésorerie', 'treso'),
        (44, 'TOTAL ACTIF', None),
    ]

    for offset, label, key in actifs:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'A{r}'].alignment = Alignment(horizontal='left')

        if turn_num == 1:
            if key:
                ws[f'C{r}'] = f'=IFERROR(INDEX({{Manufacture Belvaux,Véloce Transports,Azura Commerce,Synergia Lab},{8000,10000,4000,6000},{4000,2000,8000,2000},{2000,1000,2000,3000},{2000,3000,2000,2000}},MATCH($D$3,{{Manufacture Belvaux,Véloce Transports,Azura Commerce,Synergia Lab}},0),MATCH("{key}",{{immo,stocks,creances,treso}},0)),0)'
                # Simpler approach: use simple IF
                ws[f'C{r}'] = f'=IF($D$3="Manufacture Belvaux",IF("{key}"="immo",8000,IF("{key}"="stocks",4000,IF("{key}"="creances",2000,2000))),IF($D$3="Véloce Transports",IF("{key}"="immo",10000,IF("{key}"="stocks",2000,IF("{key}"="creances",1000,3000))),IF($D$3="Azura Commerce",IF("{key}"="immo",4000,IF("{key}"="stocks",8000,IF("{key}"="creances",2000,2000))),IF("{key}"="immo",6000,IF("{key}"="stocks",2000,IF("{key}"="creances",3000,2000))))))'
            else:
                ws[f'C{r}'] = f'=SUM(C{bs+40}:C{bs+43})'
        else:
            prev_bs = HEADER_ROWS + 1 + (turn_num - 2) * BLOCK_SIZE
            if key:
                ws[f'C{r}'] = f'=E{prev_bs+offset}'
            else:
                ws[f'C{r}'] = f'=SUM(C{bs+40}:C{bs+43})'

        if offset == 40:
            ws[f'E{r}'] = f'=C{r}+D{bs+29}'
        elif offset == 41:
            ws[f'E{r}'] = f'=C{r}+D{bs+18}+D{bs+35}'
        elif offset == 42:
            ws[f'E{r}'] = f'=C{r}+D{bs+25}'
        elif offset == 43:
            ws[f'E{r}'] = f'=C{r}+D{bs+11}+D{bs+13}+D{bs+19}+D{bs+30}+D{bs+34}'
        elif offset == 44:
            ws[f'E{r}'] = f'=SUM(E{bs+40}:E{bs+43})'

        if offset == 44:
            for col in ['A', 'C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)
        else:
            for col in ['C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_actif'])
        set_number_format(ws[f'C{r}'])
        set_number_format(ws[f'E{r}'])

    r = bs + 47
    ws[f'A{r}'] = 'PASSIF'
    ws[f'C{r}'] = 'AVANT'
    ws[f'E{r}'] = 'APRÈS'
    for col in ['A', 'C', 'E']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['light_gray'], bold=True)

    passifs = [
        (47, 'Capitaux propres', 'capitaux'),
        (48, 'Emprunts', 'emprunts'),
        (49, 'Dettes fournisseurs', 'dettes'),
        (50, 'Découvert', 'decouvert'),
        (51, 'Résultat', 'resultat'),
        (52, 'TOTAL PASSIF', None),
    ]

    for offset, label, key in passifs:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'A{r}'].alignment = Alignment(horizontal='left')

        if turn_num == 1:
            if key:
                ws[f'C{r}'] = f'=IF($D$3="Manufacture Belvaux",IF("{key}"="capitaux",8000,IF("{key}"="emprunts",8000,IF("{key}"="dettes",0,IF("{key}"="decouvert",0,0)))),IF($D$3="Véloce Transports",IF("{key}"="capitaux",8000,IF("{key}"="emprunts",8000,IF("{key}"="dettes",0,IF("{key}"="decouvert",0,0)))),IF($D$3="Azura Commerce",IF("{key}"="capitaux",8000,IF("{key}"="emprunts",8000,IF("{key}"="dettes",0,IF("{key}"="decouvert",0,0)))),IF("{key}"="capitaux",6500,IF("{key}"="emprunts",6500,IF("{key}"="dettes",0,IF("{key}"="decouvert",0,0)))))))'
            else:
                ws[f'C{r}'] = f'=SUM(C{bs+47}:C{bs+51})'
        else:
            prev_bs = HEADER_ROWS + 1 + (turn_num - 2) * BLOCK_SIZE
            if key:
                ws[f'C{r}'] = f'=E{prev_bs+offset}'
            else:
                ws[f'C{r}'] = f'=SUM(C{bs+47}:C{bs+51})'

        if offset == 47:
            ws[f'E{r}'] = f'=C{r}+D{bs+25}'
        elif offset == 48:
            ws[f'E{r}'] = f'=C{r}+D{bs+3}+D{bs+13}'
        elif offset == 49:
            ws[f'E{r}'] = f'=C{r}+D{bs+19}'
        elif offset == 50:
            ws[f'E{r}'] = f'=C{r}+D{bs+14}'
        elif offset == 51:
            ws[f'E{r}'] = f'=E{bs+58}+E{bs+64}' if turn_num == 1 else f'=C{r}+D{bs+65}'
        elif offset == 52:
            ws[f'E{r}'] = f'=SUM(E{bs+47}:E{bs+51})'

        if offset == 52:
            for col in ['A', 'C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)
        else:
            for col in ['C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_passif'])
        set_number_format(ws[f'C{r}'])
        set_number_format(ws[f'E{r}'])

    r = bs + 53
    ws[f'A{r}'] = 'Équilibre (APRÈS)'
    ws[f'C{r}'] = f'=E{bs+44}-E{bs+52}'
    set_number_format(ws[f'C{r}'])
    format_cell(ws[f'C{r}'], bg_color=COLORS['result_yellow'], bold=True)

    r = bs + 54
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = f'📈 COMPTE DE RÉSULTAT — TRIMESTRE {turn_num}'
    format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    r = bs + 55
    ws[f'A{r}'] = 'POSTE'
    ws[f'B{r}'] = '±'
    ws[f'D{r}'] = 'MONTANT'
    ws[f'F{r}'] = 'NOTE'
    for col in ['A', 'B', 'D', 'F']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['light_gray'], bold=True)

    cr_produits = [
        (56, 'CA ventes', '(+)', 'Unités × 2 000 €/u'),
        (57, 'Produits except.', '(+)', 'Subvention, remb. assurance'),
    ]

    for offset, label, sign, note in cr_produits:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'B{r}'] = sign
        ws[f'F{r}'] = note

        if offset == 56:
            ws[f'D{r}'] = f'=D{bs+25}'
        else:
            ws[f'D{r}'] = f'=IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,14,0),0)+IFERROR(VLOOKUP($D${bs+8},Données!$B$45:$L$57,8,0),0)'

        set_number_format(ws[f'D{r}'])
        for col in ['A', 'B', 'D', 'F']:
            format_cell(ws[f'{col}{r}'], bg_color=COLORS['cr_produit'])

    r = bs + 58
    ws[f'A{r}'] = 'TOTAL PRODUITS'
    ws[f'B{r}'] = 'Σ'
    ws[f'D{r}'] = f'=D{bs+56}+D{bs+57}'
    set_number_format(ws[f'D{r}'])
    for col in ['A', 'B', 'D']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)

    cr_charges = [
        (59, 'CMV', '(−)', 'Stocks à valeur d\'achat'),
        (60, 'Ch. personnel', '(−)', 'Salaires commerciaux'),
        (61, 'Ch. fixes & serv.', '(−)', 'Loyer -2000 + cartes'),
        (62, 'Dotations amort.', '(−)', 'Amortissements'),
        (63, 'Ch. intérêt', '(−)', 'Agios + intérêts emprunt'),
    ]

    for offset, label, sign, note in cr_charges:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'B{r}'] = sign
        ws[f'F{r}'] = note

        if offset == 59:
            ws[f'D{r}'] = f'=D{bs+26}'
        elif offset == 60:
            ws[f'D{r}'] = f'=IFERROR(VLOOKUP($D${bs+6},Données!$A$2:$V$38,9,0),0)+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,9,0),0)'
        elif offset == 61:
            ws[f'D{r}'] = f'=-2000+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,10,0),0)+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,11,0),0)'
        elif offset == 62:
            ws[f'D{r}'] = f'=D{bs+12}'
        elif offset == 63:
            ws[f'D{r}'] = f'=D{bs+14}+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,12,0),0)'

        set_number_format(ws[f'D{r}'])
        for col in ['A', 'B', 'D', 'F']:
            format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_passif'])

    r = bs + 64
    ws[f'A{r}'] = 'TOTAL CHARGES'
    ws[f'B{r}'] = 'Σ'
    ws[f'D{r}'] = f'=SUM(D{bs+59}:D{bs+63})'
    set_number_format(ws[f'D{r}'])
    for col in ['A', 'B', 'D']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)

    r = bs + 65
    ws[f'A{r}'] = 'RÉSULTAT NET'
    ws[f'B{r}'] = '='
    ws[f'D{r}'] = f'=D{bs+58}+D{bs+64}'
    set_number_format(ws[f'D{r}'])
    for col in ['A', 'B', 'D']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['result_yellow'], bold=True)

def main():
    xlsx_path = Path('/sessions/sharp-dazzling-babbage/mnt/jedevienspatron-github/jeu_interactif_v2.xlsx')

    print("Reading existing Données sheet...")
    donnees_data = read_donnees_sheet(str(xlsx_path))
    if donnees_data is None:
        print("FAILED: Could not read Données sheet")
        sys.exit(1)
    print(f"  Found {len(donnees_data)} cells with data")

    print("Creating new workbook...")
    wb = Workbook()
    ws_sim = wb.active
    ws_sim.title = 'Simulation'
    ws_donnees = wb.create_sheet('Données')

    print("Copying Données sheet data...")
    copy_donnees_to_new(wb, donnees_data)

    print("Building Simulation sheet...")
    build_simulation_sheet(ws_sim)

    print(f"Saving to {xlsx_path}...")
    wb.save(str(xlsx_path))
    print("SUCCESS: File created")

if __name__ == '__main__':
    main()
