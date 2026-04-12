#!/usr/bin/env python3
"""
Rebuild jeu_interactif_v2.xlsx with BLOCK_SIZE=67 and CR section.
Version 4: Add placeholder 0 values to effect cells.

BLOCK_SIZE=67:
- Offsets 0-53: existing content (turn header, inputs, effets, bilan)
- Offsets 54-65: NEW CR section
- Offset 66: blank spacer
"""

import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_ROWS = 11
BLOCK_SIZE = 67
MAX_TURNS = 12

COLORS = {
    'dark_blue': '1F3864',
    'light_blue': '2E75B6',
    'orange': 'ED7D31',
    'green': '70AD47',
    'dark_purple': '4472C4',
    'purple': '7030A0',
    'red': 'FF0000',
    'light_gray': 'D9D9D9',
    'bilan_actif': 'DEEAF1',
    'bilan_passif': 'FCE4D6',
    'bilan_total': 'BDD7EE',
    'result_yellow': 'FFF2CC',
    'cr_produit': 'E2EFDA',
    'input_yellow': 'FFFF00',
}

VALIDATION_LISTS = {
    'enterprise': ['Manufacture Belvaux', 'Véloce Transports', 'Azura Commerce', 'Synergia Lab'],
    'trimestres': ['6', '8', '10', '12'],
    'emprunt': ['0', '5000', '8000', '12000', '16000', '20000'],
    'mode': ['Tréso', 'Dettes D+1'],
    'recr': ['Aucun', 'Commercial Junior', 'Commercial Senior', 'Directrice Commerciale'],
    'carte': ['Aucune', 'Camionnette', 'Berline', 'Site Internet', 'RSE', 'R&D', 'Expansion',
              'Certification ISO', 'Application Mobile', 'Prêt Bancaire', 'Levée de Fonds',
              'Crédit-Bail', 'Crowdfunding', 'Publicité', 'Relance Clients', 'Formation',
              'Programme Fidélité', 'Export International', 'Partenariat Commercial', 'Affacturage',
              'Contrat Maintenance', 'Assurance Prévoyance', 'Mutuelle Collective', 'Cybersécurité',
              'Fourgon Réfrigéré', 'Vélo Cargo', 'ERP', 'Marketplace', 'Entrepôt Automatisé',
              'Label Qualité', 'Achat d\'Urgence', 'Maintenance Préventive', 'Révision Générale',
              'Optimisation Lean', 'Sous-traitance'],
    'event': ['Aucun', 'Client VIP', 'Contrôle Fiscal', 'Subvention Innovation', 'Placement Financier',
              'Crise Sanitaire', 'Incendie', 'Grève', 'Bouche à Oreille', 'Perte de Données',
              'Développement Durable', 'Prix PME', 'Rupture Fournisseur', 'Litige Commercial'],
}

def read_donnees_sheet(xlsx_path):
    """Read Données sheet from existing file."""
    try:
        wb = load_workbook(xlsx_path)
        if 'Données' not in wb.sheetnames:
            print("ERROR: Données sheet not found")
            return None
        sheet = wb['Données']
        data = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    data[(cell.row, cell.column)] = cell.value
        return data
    except Exception as e:
        print(f"ERROR reading Données: {e}")
        return None

def copy_donnees_to_new(new_wb, donnees_data):
    """Copy Données sheet data to new workbook."""
    sheet = new_wb['Données']
    for (row, col), value in donnees_data.items():
        sheet.cell(row, col).value = value

def add_dv(sheet, cell_ref, formula):
    """Add data validation to a cell."""
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.add(cell_ref)
    sheet.add_data_validation(dv)

def format_cell(cell, bg_color=None, font_color=None, bold=False):
    """Apply formatting to a cell."""
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    if font_color or bold:
        cell.font = Font(color=font_color or '000000', bold=bold, size=10, name='Arial')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

def set_number_format(cell):
    """Set currency format for a cell."""
    cell.number_format = '#,##0;-#,##0;"-"'

def build_simulation_sheet(ws):
    """Build the Simulation sheet with all turns."""

    # Set column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 7

    # Header section (rows 1-11)
    ws['A1'] = 'SIMULATION JEU INTERACTIF'
    ws['A1'].font = Font(bold=True, size=14, name='Arial')

    ws['A3'] = 'Entreprise :'
    ws['D3'] = 'Manufacture Belvaux'
    format_cell(ws['D3'], bg_color=COLORS['input_yellow'], font_color='0000FF')
    add_dv(ws, 'D3', f'"' + ','.join(VALIDATION_LISTS['enterprise']) + '"')

    ws['A4'] = 'Nombre de trimestres :'
    ws['D4'] = 6
    format_cell(ws['D4'], bg_color=COLORS['input_yellow'], font_color='0000FF')
    add_dv(ws, 'D4', f'"' + ','.join(VALIDATION_LISTS['trimestres']) + '"')

    ws['A5'] = 'Trimestre initial :'
    ws['D5'] = 1

    # Header row for turn blocks
    ws['A11'] = 'TRIMESTRE'
    ws['B11'] = 'AVANT'
    ws['C11'] = 'DÉCISIONS'
    ws['D11'] = 'MONTANT'
    ws['E11'] = 'APRÈS'
    ws['F11'] = 'NOTE'

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = ws[f'{col}11']
        format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    # Build turns 1-12
    for turn_num in range(1, MAX_TURNS + 1):
        build_turn_block(ws, turn_num)

def build_turn_block(ws, turn_num):
    """Build a single turn block (BLOCK_SIZE=67 rows)."""
    bs = HEADER_ROWS + 1 + (turn_num - 1) * BLOCK_SIZE  # Block start (Excel 1-indexed)

    # Turn header (row bs + 0)
    r = bs
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = f'TRIMESTRE {turn_num}'
    format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    # Décisions section header (row bs + 1)
    r = bs + 1
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = '📋 DÉCISIONS'
    format_cell(cell, bg_color=COLORS['light_blue'], font_color='FFFFFF', bold=True)

    # Decision rows (offsets 3-8)
    decisions = [
        (3, 'Emprunt demandé :', 0, VALIDATION_LISTS['emprunt']),
        (4, 'Achat de stock :', 0, None),
        (5, 'Mode de trésorerie :', 'Tréso', VALIDATION_LISTS['mode']),
        (6, 'Recrutement :', 'Aucun', VALIDATION_LISTS['recr']),
        (7, 'Carte action :', 'Aucune', VALIDATION_LISTS['carte']),
        (8, 'Événement :', 'Aucun', VALIDATION_LISTS['event']),
    ]

    for offset, label, default_val, dv_list in decisions:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'A{r}'].alignment = Alignment(horizontal='left')

        cell = ws[f'D{r}']
        cell.value = default_val
        format_cell(cell, bg_color=COLORS['input_yellow'], font_color='0000FF')
        if isinstance(default_val, (int, float)):
            set_number_format(cell)

        if dv_list:
            add_dv(ws, f'D{r}', f'"' + ','.join(dv_list) + '"')

    # Effets automatiques section header (row bs + 10)
    r = bs + 10
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = '⚙️ EFFETS AUTOMATIQUES'
    format_cell(cell, bg_color=COLORS['orange'], font_color='FFFFFF', bold=True)

    # Effets section (offsets 11-35) - initialize with 0
    effets = [
        (11, 'Charges courantes'),
        (12, 'Amortissements'),
        (13, 'Remboursements'),
        (14, 'Agios'),
        (15, 'Règlement dettes'),
        (18, 'Stock acheté'),
        (19, 'Stock à payer'),
        (22, 'Capacité prod.'),
        (23, 'Stocks dispo'),
        (24, 'Unités vendues'),
        (25, 'CA (Chiffre affaires)'),
        (26, 'CMV (Coût matière)'),
        (29, 'Immobilisations'),
        (30, 'Trésorerie'),
        (31, 'Stocks'),
        (34, 'Trésorerie'),
        (35, 'Stocks'),
    ]

    for offset, label in effets:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'A{r}'].alignment = Alignment(horizontal='left')
        ws[f'D{r}'] = 0
        set_number_format(ws[f'D{r}'])

    # Bilan section header (row bs + 39)
    r = bs + 39
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = '📊 BILAN'
    format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    # Bilan - Actif header (row bs + 40)
    r = bs + 40
    ws[f'A{r}'] = 'ACTIF'
    ws[f'C{r}'] = 'AVANT'
    ws[f'E{r}'] = 'APRÈS'
    for col in ['A', 'C', 'E']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['light_gray'], bold=True)

    # Actif rows
    actifs = [
        (40, 'Immobilisations'),
        (41, 'Stocks'),
        (42, 'Créances clients'),
        (43, 'Trésorerie'),
        (44, 'TOTAL ACTIF'),
    ]

    for offset, label in actifs:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'A{r}'].alignment = Alignment(horizontal='left')

        # Set AVANT (column C)
        if turn_num == 1:
            if offset == 40:
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,2,0),0)'
            elif offset == 41:
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,3,0),0)'
            elif offset == 42:
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,4,0),0)'
            elif offset == 43:
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,5,0),0)'
            elif offset == 44:
                ws[f'C{r}'] = f'=SUM(C{bs+40}:C{bs+43})'
        else:
            prev_bs = HEADER_ROWS + 1 + (turn_num - 2) * BLOCK_SIZE
            if offset == 40:
                ws[f'C{r}'] = f'=E{prev_bs+40}'
            elif offset == 41:
                ws[f'C{r}'] = f'=E{prev_bs+41}'
            elif offset == 42:
                ws[f'C{r}'] = f'=E{prev_bs+42}'
            elif offset == 43:
                ws[f'C{r}'] = f'=E{prev_bs+43}'
            elif offset == 44:
                ws[f'C{r}'] = f'=SUM(C{bs+40}:C{bs+43})'

        # Set APRÈS (column E)
        if offset == 40:
            ws[f'E{r}'] = f'=C{r}+D{bs+29}'
        elif offset == 41:
            ws[f'E{r}'] = f'=C{r}+D{bs+18}+D{bs+35}'
        elif offset == 42:
            ws[f'E{r}'] = f'=C{r}+D{bs+25}'
        elif offset == 43:
            ws[f'E{r}'] = f'=C{r}+D{bs+11}+D{bs+13}+D{bs+19}+D{bs+30}+D{bs+34}'
        elif offset == 44:
            ws[f'E{r}'] = f'=SUM(E{bs+40}:E{bs+43})'

        # Formatting
        if offset == 44:
            for col in ['A', 'C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)
        else:
            for col in ['C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_actif'])
        set_number_format(ws[f'C{r}'])
        set_number_format(ws[f'E{r}'])

    # Bilan - Passif section (row bs + 47)
    r = bs + 47
    ws[f'A{r}'] = 'PASSIF'
    ws[f'C{r}'] = 'AVANT'
    ws[f'E{r}'] = 'APRÈS'
    for col in ['A', 'C', 'E']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['light_gray'], bold=True)

    # Passif rows
    passifs = [
        (47, 'Capitaux propres'),
        (48, 'Emprunts'),
        (49, 'Dettes fournisseurs'),
        (50, 'Découvert'),
        (51, 'Résultat'),
        (52, 'TOTAL PASSIF'),
    ]

    for offset, label in passifs:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'A{r}'].alignment = Alignment(horizontal='left')

        # Set AVANT (column C)
        if turn_num == 1:
            if offset == 47:
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,6,0),0)'
            elif offset == 48:
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,7,0),0)'
            elif offset == 49:
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,8,0),0)'
            elif offset == 50:
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,9,0),0)'
            elif offset == 51:
                ws[f'C{r}'] = '0'
            elif offset == 52:
                ws[f'C{r}'] = f'=SUM(C{bs+47}:C{bs+51})'
        else:
            prev_bs = HEADER_ROWS + 1 + (turn_num - 2) * BLOCK_SIZE
            if offset == 47:
                ws[f'C{r}'] = f'=E{prev_bs+47}'
            elif offset == 48:
                ws[f'C{r}'] = f'=E{prev_bs+48}'
            elif offset == 49:
                ws[f'C{r}'] = f'=E{prev_bs+49}'
            elif offset == 50:
                ws[f'C{r}'] = f'=E{prev_bs+50}'
            elif offset == 51:
                ws[f'C{r}'] = f'=E{prev_bs+51}'
            elif offset == 52:
                ws[f'C{r}'] = f'=SUM(C{bs+47}:C{bs+51})'

        # Set APRÈS (column E)
        if offset == 47:
            ws[f'E{r}'] = f'=C{r}+D{bs+25}'
        elif offset == 48:
            ws[f'E{r}'] = f'=C{r}+D{bs+3}+D{bs+13}'
        elif offset == 49:
            ws[f'E{r}'] = f'=C{r}+D{bs+19}'
        elif offset == 50:
            ws[f'E{r}'] = f'=C{r}+D{bs+14}'
        elif offset == 51:
            ws[f'E{r}'] = f'=E{bs+58}+E{bs+64}' if turn_num == 1 else f'=C{r}+D{bs+65}'
        elif offset == 52:
            ws[f'E{r}'] = f'=SUM(E{bs+47}:E{bs+51})'

        # Formatting
        if offset == 52:
            for col in ['A', 'C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)
        else:
            for col in ['C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_passif'])
        set_number_format(ws[f'C{r}'])
        set_number_format(ws[f'E{r}'])

    # Équilibre (row bs + 53)
    r = bs + 53
    ws[f'A{r}'] = 'Équilibre (APRÈS)'
    ws[f'C{r}'] = f'=E{bs+44}-E{bs+52}'
    set_number_format(ws[f'C{r}'])
    format_cell(ws[f'C{r}'], bg_color=COLORS['result_yellow'], bold=True)

    # CR section header (row bs + 54)
    r = bs + 54
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = f'📈 COMPTE DE RÉSULTAT — TRIMESTRE {turn_num}'
    format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    # CR sub-header (row bs + 55)
    r = bs + 55
    ws[f'A{r}'] = 'POSTE'
    ws[f'B{r}'] = '±'
    ws[f'D{r}'] = 'MONTANT'
    ws[f'F{r}'] = 'NOTE'
    for col in ['A', 'B', 'D', 'F']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['light_gray'], bold=True)

    # CR Produits section (rows bs + 56 to bs + 58)
    cr_produits = [
        (56, 'CA ventes', '(+)', 'Unités × 2 000 €/u'),
        (57, 'Produits except.', '(+)', 'Subvention, remb. assurance'),
    ]

    for offset, label, sign, note in cr_produits:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'B{r}'] = sign
        ws[f'F{r}'] = note

        if offset == 56:
            ws[f'D{r}'] = f'=D{bs+25}'
        else:
            ws[f'D{r}'] = f'=IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,14,0),0)+IFERROR(VLOOKUP($D${bs+8},Données!$B$45:$L$57,8,0),0)'

        set_number_format(ws[f'D{r}'])
        for col in ['A', 'B', 'D', 'F']:
            format_cell(ws[f'{col}{r}'], bg_color=COLORS['cr_produit'])

    # Total produits (row bs + 58)
    r = bs + 58
    ws[f'A{r}'] = 'TOTAL PRODUITS'
    ws[f'B{r}'] = 'Σ'
    ws[f'D{r}'] = f'=D{bs+56}+D{bs+57}'
    set_number_format(ws[f'D{r}'])
    for col in ['A', 'B', 'D']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)

    # CR Charges section (rows bs + 59 to bs + 63)
    cr_charges = [
        (59, 'CMV', '(−)', 'Stocks à valeur d\'achat'),
        (60, 'Ch. personnel', '(−)', 'Salaires commerciaux'),
        (61, 'Ch. fixes & serv.', '(−)', 'Loyer -2000 + cartes'),
        (62, 'Dotations amort.', '(−)', 'Amortissements'),
        (63, 'Ch. intérêt', '(−)', 'Agios + intérêts emprunt'),
    ]

    for offset, label, sign, note in cr_charges:
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'B{r}'] = sign
        ws[f'F{r}'] = note

        if offset == 59:
            ws[f'D{r}'] = f'=D{bs+26}'
        elif offset == 60:
            ws[f'D{r}'] = f'=IFERROR(VLOOKUP($D${bs+6},Données!$A$2:$V$38,9,0),0)+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,9,0),0)'
        elif offset == 61:
            ws[f'D{r}'] = f'=-2000+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,10,0),0)+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,11,0),0)'
        elif offset == 62:
            ws[f'D{r}'] = f'=D{bs+12}'
        elif offset == 63:
            ws[f'D{r}'] = f'=D{bs+14}+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,12,0),0)'

        set_number_format(ws[f'D{r}'])
        for col in ['A', 'B', 'D', 'F']:
            format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_passif'])

    # Total charges (row bs + 64)
    r = bs + 64
    ws[f'A{r}'] = 'TOTAL CHARGES'
    ws[f'B{r}'] = 'Σ'
    ws[f'D{r}'] = f'=SUM(D{bs+59}:D{bs+63})'
    set_number_format(ws[f'D{r}'])
    for col in ['A', 'B', 'D']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)

    # Résultat net (row bs + 65)
    r = bs + 65
    ws[f'A{r}'] = 'RÉSULTAT NET'
    ws[f'B{r}'] = '='
    ws[f'D{r}'] = f'=D{bs+58}+D{bs+64}'
    set_number_format(ws[f'D{r}'])
    for col in ['A', 'B', 'D']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['result_yellow'], bold=True)

    # Spacer (row bs + 66) - empty

def main():
    xlsx_path = Path('/sessions/sharp-dazzling-babbage/mnt/jedevienspatron-github/jeu_interactif_v2.xlsx')

    print("Reading existing Données sheet...")
    donnees_data = read_donnees_sheet(str(xlsx_path))
    if donnees_data is None:
        print("FAILED: Could not read Données sheet")
        sys.exit(1)
    print(f"  Found {len(donnees_data)} cells with data")

    print("Creating new workbook...")
    wb = Workbook()
    ws_sim = wb.active
    ws_sim.title = 'Simulation'

    # Create Données sheet
    ws_donnees = wb.create_sheet('Données')

    print("Copying Données sheet data...")
    copy_donnees_to_new(wb, donnees_data)

    print("Building Simulation sheet...")
    build_simulation_sheet(ws_sim)

    print(f"Saving to {xlsx_path}...")
    wb.save(str(xlsx_path))
    print("SUCCESS: File created")

if __name__ == '__main__':
    main()
