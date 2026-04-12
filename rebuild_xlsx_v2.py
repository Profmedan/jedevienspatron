#!/usr/bin/env python3
"""
Rebuild jeu_interactif_v2.xlsx with BLOCK_SIZE=67 and CR section.
Version 2: Fix all formulas and add AVANT references.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy as copy_cell_style
import pandas as pd

HEADER_ROWS = 11
BLOCK_SIZE = 67
MAX_TURNS = 12

OFFSETS = {
    'turn_header': 0,
    'input_emprunt': 3,
    'input_achat': 4,
    'input_mode': 5,
    'input_recr': 6,
    'input_carte': 7,
    'input_event': 8,
    'eff_charges': 11,
    'eff_amort': 12,
    'eff_remb': 13,
    'eff_agios': 14,
    'eff_dettes': 15,
    'eff_achat_stk': 18,
    'eff_achat_pay': 19,
    'eff_capacite': 22,
    'eff_stocks_dispo': 23,
    'eff_units': 24,
    'eff_ca': 25,
    'eff_cmv': 26,
    'eff_carte_immo': 29,
    'eff_carte_treso': 30,
    'eff_carte_stk': 31,
    'eff_event_treso': 34,
    'eff_event_stk': 35,
    'bilan_immo': 40,
    'bilan_stocks': 41,
    'bilan_creances': 42,
    'bilan_treso': 43,
    'bilan_total_actif': 44,
    'bilan_capitaux': 47,
    'bilan_emprunts': 48,
    'bilan_dettes': 49,
    'bilan_decouvert': 50,
    'bilan_resultat': 51,
    'bilan_total_passif': 52,
    'bilan_equilibre': 53,
    'cr_header': 54,
    'cr_subhdr': 55,
    'cr_ca': 56,
    'cr_prodexcept': 57,
    'cr_totprod': 58,
    'cr_cmv': 59,
    'cr_charpers': 60,
    'cr_charfixes': 61,
    'cr_dotations': 62,
    'cr_charint': 63,
    'cr_totchg': 64,
    'cr_resultat': 65,
    'spacer': 66,
}

COLORS = {
    'dark_blue': '1F3864',
    'light_blue': '2E75B6',
    'orange': 'ED7D31',
    'green': '70AD47',
    'dark_purple': '4472C4',
    'purple': '7030A0',
    'red': 'FF0000',
    'light_gray': 'D9D9D9',
    'bilan_actif': 'DEEAF1',
    'bilan_passif': 'FCE4D6',
    'bilan_total': 'BDD7EE',
    'result_yellow': 'FFF2CC',
    'cr_produit': 'E2EFDA',
    'input_yellow': 'FFFF00',
}

VALIDATION_LISTS = {
    'enterprise': ['Manufacture Belvaux', 'Véloce Transports', 'Azura Commerce', 'Synergia Lab'],
    'trimestres': ['6', '8', '10', '12'],
    'emprunt': ['0', '5000', '8000', '12000', '16000', '20000'],
    'mode': ['Tréso', 'Dettes D+1'],
    'recr': ['Aucun', 'Commercial Junior', 'Commercial Senior', 'Directrice Commerciale'],
    'carte': ['Aucune', 'Camionnette', 'Berline', 'Site Internet', 'RSE', 'R&D', 'Expansion',
              'Certification ISO', 'Application Mobile', 'Prêt Bancaire', 'Levée de Fonds',
              'Crédit-Bail', 'Crowdfunding', 'Publicité', 'Relance Clients', 'Formation',
              'Programme Fidélité', 'Export International', 'Partenariat Commercial', 'Affacturage',
              'Contrat Maintenance', 'Assurance Prévoyance', 'Mutuelle Collective', 'Cybersécurité',
              'Fourgon Réfrigéré', 'Vélo Cargo', 'ERP', 'Marketplace', 'Entrepôt Automatisé',
              'Label Qualité', 'Achat d\'Urgence', 'Maintenance Préventive', 'Révision Générale',
              'Optimisation Lean', 'Sous-traitance'],
    'event': ['Aucun', 'Client VIP', 'Contrôle Fiscal', 'Subvention Innovation', 'Placement Financier',
              'Crise Sanitaire', 'Incendie', 'Grève', 'Bouche à Oreille', 'Perte de Données',
              'Développement Durable', 'Prix PME', 'Rupture Fournisseur', 'Litige Commercial'],
}

def read_donnees_sheet(xlsx_path):
    """Read Données sheet from existing file."""
    try:
        wb = load_workbook(xlsx_path)
        if 'Données' not in wb.sheetnames:
            print("ERROR: Données sheet not found")
            return None
        sheet = wb['Données']
        data = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    data[(cell.row, cell.column)] = {
                        'value': cell.value,
                    }
        return data
    except Exception as e:
        print(f"ERROR reading Données: {e}")
        return None

def copy_donnees_to_new(new_wb, donnees_data):
    """Copy Données sheet data to new workbook."""
    sheet = new_wb['Données']
    for (row, col), cell_data in donnees_data.items():
        cell = sheet.cell(row, col)
        cell.value = cell_data['value']

def add_dv(sheet, cell_ref, formula):
    """Add data validation to a cell."""
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.add(cell_ref)
    sheet.add_data_validation(dv)

def format_cell(cell, bg_color=None, font_color=None, bold=False, border=False):
    """Apply formatting to a cell."""
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    if font_color or bold:
        cell.font = Font(color=font_color or '000000', bold=bold, size=10, name='Arial')
    if border:
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

def set_number_format(cell):
    """Set currency format for a cell."""
    cell.number_format = '#,##0;-#,##0;"-"'

def build_simulation_sheet(ws, donnees_data):
    """Build the Simulation sheet with all turns."""

    # Set column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 7

    # Header rows (1-11)
    ws['A1'] = 'SIMULATION JEU INTERACTIF'
    ws['A1'].font = Font(bold=True, size=14, name='Arial')

    ws['A3'] = 'Entreprise :'
    ws['D3'] = 'Manufacture Belvaux'
    format_cell(ws['D3'], bg_color=COLORS['input_yellow'], font_color='0000FF')
    add_dv(ws, 'D3', f'"' + ','.join(VALIDATION_LISTS['enterprise']) + '"')

    ws['A4'] = 'Nombre de trimestres :'
    ws['D4'] = 6
    format_cell(ws['D4'], bg_color=COLORS['input_yellow'], font_color='0000FF')
    add_dv(ws, 'D4', f'"' + ','.join(VALIDATION_LISTS['trimestres']) + '"')

    ws['A5'] = 'Trimestre initial :'
    ws['D5'] = 1

    # Add header row for turn blocks
    ws['A11'] = 'TRIMESTRE'
    ws['B11'] = 'AVANT'
    ws['C11'] = 'DÉCISIONS'
    ws['D11'] = 'MONTANT'
    ws['E11'] = 'APRÈS'
    ws['F11'] = 'NOTE'

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = ws[f'{col}11']
        format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    # Build turns 1-12
    for turn_num in range(1, MAX_TURNS + 1):
        build_turn_block(ws, turn_num)

def build_turn_block(ws, turn_num):
    """Build a single turn block (BLOCK_SIZE=67 rows)."""
    bs = HEADER_ROWS + 1 + (turn_num - 1) * BLOCK_SIZE

    # Turn header
    r = bs + OFFSETS['turn_header']
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = f'TRIMESTRE {turn_num}'
    format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    # Décisions section header
    r = bs + 1
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = '📋 DÉCISIONS'
    format_cell(cell, bg_color=COLORS['light_blue'], font_color='FFFFFF', bold=True)

    # Decision rows (offsets 3-8)
    decisions = [
        ('input_emprunt', 'Emprunt demandé :', 0, VALIDATION_LISTS['emprunt']),
        ('input_achat', 'Achat de stock :', 0, None),
        ('input_mode', 'Mode de trésorerie :', 'Tréso', VALIDATION_LISTS['mode']),
        ('input_recr', 'Recrutement :', 'Aucun', VALIDATION_LISTS['recr']),
        ('input_carte', 'Carte action :', 'Aucune', VALIDATION_LISTS['carte']),
        ('input_event', 'Événement :', 'Aucun', VALIDATION_LISTS['event']),
    ]

    for key, label, default_val, dv_list in decisions:
        offset = OFFSETS[key]
        r = bs + offset

        ws[f'A{r}'] = label
        ws[f'A{r}'].alignment = Alignment(horizontal='left')

        cell = ws[f'D{r}']
        cell.value = default_val
        format_cell(cell, bg_color=COLORS['input_yellow'], font_color='0000FF')
        set_number_format(cell)

        if dv_list:
            add_dv(ws, f'D{r}', f'"' + ','.join(dv_list) + '"')

    # Effets automatiques section header
    r = bs + 10
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = '⚙️ EFFETS AUTOMATIQUES'
    format_cell(cell, bg_color=COLORS['orange'], font_color='FFFFFF', bold=True)

    # Effets rows (add formulas later from original file)
    effets_labels = [
        ('eff_charges', 'Charges courantes'),
        ('eff_amort', 'Amortissements'),
        ('eff_remb', 'Remboursements'),
        ('eff_agios', 'Agios'),
        ('eff_dettes', 'Règlement dettes'),
        (None, None),
        ('eff_achat_stk', 'Stock acheté'),
        ('eff_achat_pay', 'Stock à payer'),
        (None, None),
        ('eff_capacite', 'Capacité prod.'),
        ('eff_stocks_dispo', 'Stocks dispo'),
        ('eff_units', 'Unités vendues'),
        ('eff_ca', 'CA (Chiffre affaires)'),
        ('eff_cmv', 'CMV (Coût matière)'),
        (None, None),
        ('eff_carte_immo', 'Immobilisations'),
        ('eff_carte_treso', 'Trésorerie'),
        ('eff_carte_stk', 'Stocks'),
        (None, None),
        ('eff_event_treso', 'Trésorerie'),
        ('eff_event_stk', 'Stocks'),
    ]

    for key, label in effets_labels:
        if key is None:
            continue
        offset = OFFSETS[key]
        r = bs + offset
        if label:
            ws[f'A{r}'] = label
            ws[f'A{r}'].alignment = Alignment(horizontal='left')

    # Bilan section header
    r = bs + 39
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = '📊 BILAN'
    format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    # Bilan header
    r = bs + 40
    ws[f'A{r}'] = 'ACTIF'
    ws[f'C{r}'] = 'AVANT'
    ws[f'E{r}'] = 'APRÈS'
    for col in ['A', 'C', 'E']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['light_gray'], bold=True)

    bilan_labels = [
        ('bilan_immo', 'Immobilisations'),
        ('bilan_stocks', 'Stocks'),
        ('bilan_creances', 'Créances clients'),
        ('bilan_treso', 'Trésorerie'),
        ('bilan_total_actif', 'TOTAL ACTIF'),
    ]

    for key, label in bilan_labels:
        offset = OFFSETS[key]
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'A{r}'].alignment = Alignment(horizontal='left')

        # Set AVANT references
        if turn_num == 1:
            # For turn 1, use VLOOKUP from Données
            if key == 'bilan_immo':
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,2,0),0)'
            elif key == 'bilan_stocks':
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,3,0),0)'
            elif key == 'bilan_creances':
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,4,0),0)'
            elif key == 'bilan_treso':
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,5,0),0)'
            elif key == 'bilan_total_actif':
                ws[f'C{r}'] = f'=SUM(C{bs+40}:C{bs+43})'
        else:
            # For turn n>1, reference previous turn's APRÈS
            prev_bs = HEADER_ROWS + 1 + (turn_num - 2) * BLOCK_SIZE
            if key == 'bilan_immo':
                ws[f'C{r}'] = f'=E{prev_bs+40}'
            elif key == 'bilan_stocks':
                ws[f'C{r}'] = f'=E{prev_bs+41}'
            elif key == 'bilan_creances':
                ws[f'C{r}'] = f'=E{prev_bs+42}'
            elif key == 'bilan_treso':
                ws[f'C{r}'] = f'=E{prev_bs+43}'
            elif key == 'bilan_total_actif':
                ws[f'C{r}'] = f'=SUM(C{bs+40}:C{bs+43})'

        # Set APRÈS formulas (calculations from effects)
        if key == 'bilan_immo':
            ws[f'E{r}'] = f'=C{r}+D{bs+29}'
        elif key == 'bilan_stocks':
            ws[f'E{r}'] = f'=C{r}+D{bs+18}+D{bs+35}'
        elif key == 'bilan_creances':
            ws[f'E{r}'] = f'=C{r}+D{bs+25}'
        elif key == 'bilan_treso':
            ws[f'E{r}'] = f'=C{r}+D{bs+11}+D{bs+13}+D{bs+19}+D{bs+30}+D{bs+34}'
        elif key == 'bilan_total_actif':
            ws[f'E{r}'] = f'=SUM(E{bs+40}:E{bs+43})'

        if 'total' in key:
            for col in ['A', 'C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)
            set_number_format(ws[f'C{r}'])
            set_number_format(ws[f'E{r}'])
        else:
            for col in ['C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_actif'])
            set_number_format(ws[f'C{r}'])
            set_number_format(ws[f'E{r}'])

    # Passif section
    r = bs + 47
    ws[f'A{r}'] = 'PASSIF'
    ws[f'C{r}'] = 'AVANT'
    ws[f'E{r}'] = 'APRÈS'
    for col in ['A', 'C', 'E']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['light_gray'], bold=True)

    passif_labels = [
        ('bilan_capitaux', 'Capitaux propres'),
        ('bilan_emprunts', 'Emprunts'),
        ('bilan_dettes', 'Dettes fournisseurs'),
        ('bilan_decouvert', 'Découvert'),
        ('bilan_resultat', 'Résultat'),
        ('bilan_total_passif', 'TOTAL PASSIF'),
    ]

    for key, label in passif_labels:
        offset = OFFSETS[key]
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'A{r}'].alignment = Alignment(horizontal='left')

        # Set AVANT references
        if turn_num == 1:
            if key == 'bilan_capitaux':
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,6,0),0)'
            elif key == 'bilan_emprunts':
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,7,0),0)'
            elif key == 'bilan_dettes':
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,8,0),0)'
            elif key == 'bilan_decouvert':
                ws[f'C{r}'] = '=IFERROR(VLOOKUP($D$3,Données!$A$65:$J$68,9,0),0)'
            elif key == 'bilan_resultat':
                ws[f'C{r}'] = '0'
            elif key == 'bilan_total_passif':
                ws[f'C{r}'] = f'=SUM(C{bs+47}:C{bs+51})'
        else:
            prev_bs = HEADER_ROWS + 1 + (turn_num - 2) * BLOCK_SIZE
            if key == 'bilan_capitaux':
                ws[f'C{r}'] = f'=E{prev_bs+47}'
            elif key == 'bilan_emprunts':
                ws[f'C{r}'] = f'=E{prev_bs+48}'
            elif key == 'bilan_dettes':
                ws[f'C{r}'] = f'=E{prev_bs+49}'
            elif key == 'bilan_decouvert':
                ws[f'C{r}'] = f'=E{prev_bs+50}'
            elif key == 'bilan_resultat':
                ws[f'C{r}'] = f'=E{prev_bs+51}'
            elif key == 'bilan_total_passif':
                ws[f'C{r}'] = f'=SUM(C{bs+47}:C{bs+51})'

        # Set APRÈS formulas
        if key == 'bilan_capitaux':
            ws[f'E{r}'] = f'=C{r}+D{bs+25}'
        elif key == 'bilan_emprunts':
            ws[f'E{r}'] = f'=C{r}+D{bs+3}+D{bs+13}'
        elif key == 'bilan_dettes':
            ws[f'E{r}'] = f'=C{r}+D{bs+19}'
        elif key == 'bilan_decouvert':
            ws[f'E{r}'] = f'=C{r}+D{bs+14}'
        elif key == 'bilan_resultat':
            ws[f'E{r}'] = f'=E{bs+58}+E{bs+64}' if turn_num == 1 else f'=C{r}+D{bs+65}'
        elif key == 'bilan_total_passif':
            ws[f'E{r}'] = f'=SUM(E{bs+47}:E{bs+51})'

        if 'total' in key:
            for col in ['A', 'C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)
            set_number_format(ws[f'C{r}'])
            set_number_format(ws[f'E{r}'])
        else:
            for col in ['C', 'E']:
                format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_passif'])
            set_number_format(ws[f'C{r}'])
            set_number_format(ws[f'E{r}'])

    # Équilibre
    r = bs + 53
    ws[f'A{r}'] = 'Équilibre (APRÈS)'
    ws[f'C{r}'] = f'=E{bs+44}-E{bs+52}'
    set_number_format(ws[f'C{r}'])
    format_cell(ws[f'C{r}'], bg_color=COLORS['result_yellow'], bold=True)

    # CR section header
    r = bs + OFFSETS['cr_header']
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws[f'A{r}']
    cell.value = f'📈 COMPTE DE RÉSULTAT — TRIMESTRE {turn_num}'
    format_cell(cell, bg_color=COLORS['dark_blue'], font_color='FFFFFF', bold=True)

    # CR sub-header
    r = bs + OFFSETS['cr_subhdr']
    ws[f'A{r}'] = 'POSTE'
    ws[f'B{r}'] = '±'
    ws[f'D{r}'] = 'MONTANT'
    ws[f'F{r}'] = 'NOTE'
    for col in ['A', 'B', 'D', 'F']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['light_gray'], bold=True)

    # CR produits
    cr_produits = [
        ('cr_ca', 'CA ventes', '(+)', 'Unités × 2 000 €/u'),
        ('cr_prodexcept', 'Produits except.', '(+)', 'Subvention, remb. assurance'),
    ]

    for key, label, sign, note in cr_produits:
        offset = OFFSETS[key]
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'B{r}'] = sign
        ws[f'F{r}'] = note

        if key == 'cr_ca':
            ws[f'D{r}'] = f'=D{bs+25}'
        else:
            ws[f'D{r}'] = f'=IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,14,0),0)+IFERROR(VLOOKUP($D${bs+8},Données!$B$45:$L$57,8,0),0)'

        set_number_format(ws[f'D{r}'])
        for col in ['A', 'B', 'D', 'F']:
            format_cell(ws[f'{col}{r}'], bg_color=COLORS['cr_produit'])

    # Total produits
    r = bs + OFFSETS['cr_totprod']
    ws[f'A{r}'] = 'TOTAL PRODUITS'
    ws[f'B{r}'] = 'Σ'
    ws[f'D{r}'] = f'=D{bs+56}+D{bs+57}'
    set_number_format(ws[f'D{r}'])
    for col in ['A', 'B', 'D']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)

    # CR charges
    cr_charges = [
        ('cr_cmv', 'CMV', '(−)', 'Stocks à valeur d\'achat'),
        ('cr_charpers', 'Ch. personnel', '(−)', 'Salaires commerciaux'),
        ('cr_charfixes', 'Ch. fixes & serv.', '(−)', 'Loyer -2000 + cartes'),
        ('cr_dotations', 'Dotations amort.', '(−)', 'Amortissements'),
        ('cr_charint', 'Ch. intérêt', '(−)', 'Agios + intérêts emprunt'),
    ]

    for key, label, sign, note in cr_charges:
        offset = OFFSETS[key]
        r = bs + offset
        ws[f'A{r}'] = label
        ws[f'B{r}'] = sign
        ws[f'F{r}'] = note

        if key == 'cr_cmv':
            ws[f'D{r}'] = f'=D{bs+26}'
        elif key == 'cr_charpers':
            ws[f'D{r}'] = f'=IFERROR(VLOOKUP($D${bs+6},Données!$A$2:$V$38,9,0),0)+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,9,0),0)'
        elif key == 'cr_charfixes':
            ws[f'D{r}'] = f'=-2000+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,10,0),0)+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,11,0),0)'
        elif key == 'cr_dotations':
            ws[f'D{r}'] = f'=D{bs+12}'
        elif key == 'cr_charint':
            ws[f'D{r}'] = f'=D{bs+14}+IFERROR(VLOOKUP($D${bs+7},Données!$A$2:$V$38,12,0),0)'

        set_number_format(ws[f'D{r}'])
        for col in ['A', 'B', 'D', 'F']:
            format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_passif'])

    # Total charges
    r = bs + OFFSETS['cr_totchg']
    ws[f'A{r}'] = 'TOTAL CHARGES'
    ws[f'B{r}'] = 'Σ'
    ws[f'D{r}'] = f'=SUM(D{bs+59}:D{bs+63})'
    set_number_format(ws[f'D{r}'])
    for col in ['A', 'B', 'D']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['bilan_total'], bold=True)

    # Résultat net
    r = bs + OFFSETS['cr_resultat']
    ws[f'A{r}'] = 'RÉSULTAT NET'
    ws[f'B{r}'] = '='
    ws[f'D{r}'] = f'=D{bs+58}+D{bs+64}'
    set_number_format(ws[f'D{r}'])
    for col in ['A', 'B', 'D']:
        format_cell(ws[f'{col}{r}'], bg_color=COLORS['result_yellow'], bold=True)

def main():
    xlsx_path = Path('/sessions/sharp-dazzling-babbage/mnt/jedevienspatron-github/jeu_interactif_v2.xlsx')

    print("Reading existing Données sheet...")
    donnees_data = read_donnees_sheet(str(xlsx_path))
    if donnees_data is None:
        print("FAILED: Could not read Données sheet")
        sys.exit(1)
    print(f"  Found {len(donnees_data)} cells with data")

    print("Creating new workbook...")
    wb = Workbook()
    ws_sim = wb.active
    ws_sim.title = 'Simulation'

    # Create Données sheet
    ws_donnees = wb.create_sheet('Données')

    print("Copying Données sheet data...")
    copy_donnees_to_new(wb, donnees_data)

    print("Building Simulation sheet...")
    build_simulation_sheet(ws_sim, donnees_data)

    print(f"Saving to {xlsx_path}...")
    wb.save(str(xlsx_path))
    print("SUCCESS: File created")

if __name__ == '__main__':
    main()
